from .models import Producto
import openpyxl
from django.db.models import Q
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import MovimientoInventario, Producto
from django.shortcuts import render, redirect
from django.db.models import Sum 
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import redirect, get_object_or_404
from .models import ConteoInventario,ItemConteo
from openpyxl.styles import Font
from django.core.paginator import Paginator



def login_view(request):
    if request.user.is_authenticated:
        return redirect('lista_productos')  # ya está logueado

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('lista_productos')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, "productos/login.html")

def cargar_excel(request):
    if request.method == "POST" and request.FILES.get("archivo"):
        archivo = request.FILES["archivo"]

        if not archivo.name.endswith(".xlsx"):
            messages.error(request, "El archivo debe ser .xlsx")
            return redirect("cargar_excel")

        wb = openpyxl.load_workbook(archivo)
        hoja = wb.active

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            codigo, referencia, departamento, seccion, descripcion, tipo, stock = fila

            if codigo:
                Producto.objects.update_or_create(
                    codigo_barras=codigo,
                    defaults={
                        "referencia": referencia,
                        "departamento": departamento,
                        "seccion": seccion,
                        "descripcion": descripcion,
                        "tipo": tipo,
                        "stock": stock or 0,
                    }
                )

        messages.success(request, "Archivo cargado correctamente")
        return redirect("cargar_excel")

    return render(request, "productos/cargar_excel.html")


@login_required
def lista_productos(request):
    query = request.GET.get("q", "")
    filtro_departamento = request.GET.get("departamento", "")
    filtro_seccion = request.GET.get("seccion", "")
    filtro_tipo = request.GET.get("tipo", "")

    productos = Producto.objects.all()

    if query:
        productos = productos.filter(
            Q(codigo_barras__icontains=query) |
            Q(referencia__icontains=query) |
            Q(descripcion__icontains=query)
        )
    if filtro_departamento:
        productos = productos.filter(departamento=filtro_departamento)
    if filtro_seccion:
        productos = productos.filter(seccion=filtro_seccion)
    if filtro_tipo:
        productos = productos.filter(tipo=filtro_tipo)

    # Paginación aquí (20 productos por página)
    paginator = Paginator(productos, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "productos": page_obj,  # cambia productos por page_obj
        "query": query,
        "filtro_departamento": filtro_departamento,
        "filtro_seccion": filtro_seccion,
        "filtro_tipo": filtro_tipo,
    }
    return render(request, "productos/lista_productos.html", context)


from django.contrib.auth import logout

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def historial_movimientos(request):
    movimientos = MovimientoInventario.objects.select_related('producto', 'usuario').order_by('-fecha')

    producto_id = request.GET.get('producto')
    tipo = request.GET.get('tipo')

    if producto_id:
        movimientos = movimientos.filter(producto_id=producto_id)
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)

    productos = Producto.objects.all()
    return render(request, "productos/historial_movimientos.html", {
        "movimientos": movimientos,
        "productos": productos,
        "producto_id": producto_id,
        "tipo": tipo,
    })

@login_required
def dashboard(request):
    total_productos = Producto.objects.count()
    total_stock = Producto.objects.aggregate(Sum('stock'))['stock__sum'] or 0
    stock_bajo = Producto.objects.filter(stock__lt=10)
    ultimos_movimientos = MovimientoInventario.objects.select_related('producto', 'usuario').order_by('-fecha')[:5]

    return render(request, 'productos/dashboard.html', {
        'total_productos': total_productos,
        'total_stock': total_stock,
        'stock_bajo': stock_bajo,
        'ultimos_movimientos': ultimos_movimientos,
    })

@login_required
def crear_movimiento(request):
    productos = Producto.objects.all()
    producto_obj = None
    mostrar_formulario = False

    # Cargar automáticamente si viene por GET desde productos
    if request.method == "GET" and "codigo" in request.GET:
        codigo_barras = request.GET.get("codigo")
        try:
            producto_obj = Producto.objects.get(codigo_barras=codigo_barras)
            mostrar_formulario = True
        except Producto.DoesNotExist:
            messages.error(request, "Producto no encontrado con ese código de barras.")

    elif request.method == "POST":
        if 'buscar' in request.POST:
            codigo_barras = request.POST.get("codigo_barras")
            try:
                producto_obj = Producto.objects.get(codigo_barras=codigo_barras)
                mostrar_formulario = True
            except Producto.DoesNotExist:
                messages.error(request, "Producto no encontrado con ese código de barras.")

        elif 'registrar' in request.POST:
            codigo_barras = request.POST.get("codigo_barras")
            tipo = request.POST.get("tipo")
            cantidad = int(request.POST.get("cantidad"))

            try:
                producto = Producto.objects.get(codigo_barras=codigo_barras)
                producto_obj = producto
                mostrar_formulario = True

                if tipo == "salida" and producto.stock < cantidad:
                    messages.error(request, "No hay suficiente stock para realizar esta salida.")
                    return redirect("crear_movimiento")

                if tipo == "entrada":
                    producto.stock += cantidad
                elif tipo == "salida":
                    producto.stock -= cantidad

                producto.save()

                MovimientoInventario.objects.create(
                    producto=producto,
                    tipo=tipo,
                    cantidad=cantidad,
                    usuario=request.user
                )

                messages.success(request, f"{tipo.capitalize()} registrada correctamente para {producto.descripcion}.")
                return redirect("crear_movimiento")

            except Producto.DoesNotExist:
                messages.error(request, "Producto no encontrado con ese código de barras.")

    return render(request, "productos/crear_movimiento.html", {
        "productos": productos,
        "producto_obj": producto_obj,
        "mostrar_formulario": mostrar_formulario,
    })

@csrf_exempt
@login_required
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == "POST":
        producto.descripcion = request.POST.get("descripcion", producto.descripcion)
        producto.referencia = request.POST.get("referencia", producto.referencia)
        producto.departamento = request.POST.get("departamento", producto.departamento)
        producto.seccion = request.POST.get("seccion", producto.seccion)
        producto.tipo = request.POST.get("tipo", producto.tipo)
        producto.save()

    return redirect('lista_productos') 

from django.http import FileResponse
from django.utils.timezone import now
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.graphics.barcode import code128
from reportlab.lib.units import mm
from .models import Producto
import io

def generar_etiqueta(request, producto_id):
    producto = Producto.objects.get(id=producto_id)
    fecha = now().strftime("%d/%m/%Y")

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    # Título y fecha
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, 770, "Inventario Adaros")
    c.setFont("Helvetica", 10)
    c.drawString(50, 755, f"Fecha: {fecha}")

    # Información del producto
    c.setFont("Helvetica", 11)
    c.drawString(50, 730, f"Producto: {producto.descripcion[:50]}")
    c.drawString(50, 715, f"Código: {producto.codigo_barras}")

    # Generar código de barras
    barcode = code128.Code128(str(producto.codigo_barras), barHeight=20*mm, barWidth=0.4)
    barcode.drawOn(c, 50, 650)

    c.showPage()
    c.save()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename=f"etiqueta_{producto.codigo_barras}.pdf")


def buscar_producto(request):
    query = request.GET.get("q", "")
    resultados = Producto.objects.filter(
        descripcion__icontains=query
    ) | Producto.objects.filter(codigo_barras__icontains=query)

    return render(request, "productos/busqueda.html", {
        "query": query,
        "resultados": resultados,
    })

from django.utils.timezone import now, timedelta

def detalle_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    ultimos_movimientos = (
        MovimientoInventario.objects
        .filter(producto=producto)
        .order_by("-fecha")[:10]
    )

    # Agrupar por fecha
    datos = {
        "fechas": [],
        "entradas": [],
        "salidas": [],
    }

    for mov in reversed(ultimos_movimientos):
        fecha_str = mov.fecha.strftime("%d/%m")
        datos["fechas"].append(fecha_str)
        datos["entradas"].append(mov.cantidad if mov.tipo == "entrada" else 0)
        datos["salidas"].append(mov.cantidad if mov.tipo == "salida" else 0)

    return render(request, "productos/detalle_producto.html", {
        "producto": producto,
        "datos": datos
    })


@login_required
def crear_conteo(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        if nombre:
            ConteoInventario.objects.create(nombre=nombre, creado_por=request.user)
            messages.success(request, "Sesión de conteo creada correctamente.")
            return redirect("lista_conteos")
        else:
            messages.error(request, "Debes ingresar un nombre para el conteo.")
    return render(request, "conteo/crear_conteo.html")

@login_required
def lista_conteos(request):
    conteos = ConteoInventario.objects.all().order_by('-fecha_creado')
    return render(request, "conteo/lista_conteos.html", {
        "conteos": conteos
    })


@login_required
def detalle_conteo(request, conteo_id):
    conteo = get_object_or_404(ConteoInventario, id=conteo_id)
    producto_encontrado = None

    if request.method == "POST" and not conteo.cerrado:
        codigo = request.POST.get("codigo_barras")
        cantidad = request.POST.get("cantidad")

        try:
            producto = Producto.objects.get(codigo_barras=codigo)
            producto_encontrado = producto

            if cantidad and cantidad.isdigit():
                ItemConteo.objects.create(
                    conteo=conteo,
                    producto=producto,
                    cantidad_contada=int(cantidad),
                    usuario=request.user
                )
                messages.success(request, f"{producto.descripcion} agregado con cantidad {cantidad}.")
                return redirect("detalle_conteo", conteo_id=conteo.id)
            else:
                messages.error(request, "Cantidad inválida.")
        except Producto.DoesNotExist:
            messages.error(request, "Producto no encontrado.")

    items = conteo.items.select_related("producto").order_by("-fecha")

    conteo_totales_raw = (
        ItemConteo.objects
        .filter(conteo=conteo)
        .values('producto__id', 'producto__descripcion', 'producto__stock')
        .annotate(total_contado=Sum('cantidad_contada'))
        .order_by('producto__descripcion')
    )

    conteo_totales = []
    for item in conteo_totales_raw:
        diferencia = item["total_contado"] - item["producto__stock"]
        conteo_totales.append({
            "descripcion": item["producto__descripcion"],
            "stock": item["producto__stock"],
            "contado": item["total_contado"],
            "diferencia": diferencia
        })

    return render(request, "conteo/detalle_conteo.html", {
        "conteo": conteo,
        "items": items,
        "producto_encontrado": producto_encontrado,
        "conteo_totales": conteo_totales
    })
                                               

from django.views.decorators.http import require_POST

@require_POST
@login_required
def cerrar_conteo(request, conteo_id):
    conteo = get_object_or_404(ConteoInventario, id=conteo_id)
    if not conteo.cerrado:
        conteo.cerrado = True
        conteo.save()
        messages.success(request, "Conteo cerrado correctamente.")
    return redirect("detalle_conteo", conteo_id=conteo.id)

@require_POST
@login_required
def ajustar_stock_conteo(request, conteo_id):
    conteo = get_object_or_404(ConteoInventario, id=conteo_id)

    if not conteo.cerrado:
        messages.error(request, "El conteo debe estar cerrado antes de ajustar el stock.")
        return redirect("detalle_conteo", conteo_id=conteo.id)

    ajustes = {}
    items = ItemConteo.objects.filter(conteo=conteo).values("producto").annotate(
        total_contado=Sum("cantidad_contada")
    )

    for item in items:
        producto = Producto.objects.get(id=item["producto"])
        nuevo_stock = item["total_contado"]
        ajustes[producto.codigo_barras] = (producto.stock, nuevo_stock)
        producto.stock = nuevo_stock
        producto.save()

    messages.success(request, f"Stock ajustado para {len(ajustes)} productos.")
    return redirect("detalle_conteo", conteo_id=conteo.id)


from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from django.db.models import Sum
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import ConteoInventario, ItemConteo
from reportlab.platypus import Image
import os
from django.conf import settings

def generar_reporte_conteo_pdf(request, conteo_id):
    conteo = get_object_or_404(ConteoInventario, id=conteo_id)
    if not conteo.cerrado:
        messages.error(request, "Debes cerrar el conteo antes de generar el reporte.")
        return redirect("detalle_conteo", conteo_id=conteo.id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_conteo_{conteo.id}.pdf"'

    logo_path = os.path.join(settings.BASE_DIR, 'static', 'productos', 'adaros_asesores.png')
    try:
        if os.path.exists(logo_path):
            p.drawImage(logo_path, 40, height - 100, width=120, preserveAspectRatio=True, mask='auto')
    except Exception as e:
        print(f"Error al cargar el logo: {e}")

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50

    # Encabezado institucional
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(width / 2, y, "ADARO ASESORES")
    y -= 18

    p.setFont("Helvetica", 10)
    p.drawCentredString(width / 2, y, "Servicios Administrativos y de Inventario")
    y -= 12
    p.drawCentredString(width / 2, y, "Tel: +1 (849) 720-6153 | Correo: _______________________")
    y -= 25

    # Título del documento
    p.setFont("Helvetica-Bold", 13)
    p.drawCentredString(width / 2, y, "ACTA DE CONTEO FÍSICO DE INVENTARIO")
    y -= 30

    # Información del conteo
    p.setFont("Helvetica", 10)
    p.drawString(40, y, f"Nombre del conteo: {conteo.nombre}")
    y -= 15
    p.drawString(40, y, f"Fecha de realización: {conteo.fecha_creado.strftime('%d/%m/%Y %H:%M')}")
    y -= 15
    p.drawString(40, y, f"Responsable del conteo: {conteo.creado_por}")
    y -= 25

    # Cabecera de tabla
    p.setFont("Helvetica-Bold", 10)
    p.drawString(40, y, "Producto")
    p.drawString(240, y, "Stock previo")
    p.drawString(320, y, "Contado")
    p.drawString(400, y, "Diferencia")
    y -= 15

    p.setFont("Helvetica", 9)

    totales = (
        ItemConteo.objects
        .filter(conteo=conteo)
        .values('producto__descripcion', 'producto__stock')
        .annotate(total_contado=Sum('cantidad_contada'))
        .order_by('producto__descripcion')
    )

    for item in totales:
        if y < 100:
            p.showPage()
            y = height - 50
        desc = item["producto__descripcion"][:40]
        stock = item["producto__stock"]
        contado = item["total_contado"]
        diferencia = contado - stock

        p.drawString(40, y, desc)
        p.drawRightString(280, y, str(stock))
        p.drawRightString(360, y, str(contado))
        p.drawRightString(440, y, str(diferencia))
        y -= 15

    # Área de firmas
    y -= 50
    p.setFont("Helvetica", 10)
    p.drawString(80, y, "______________________________")
    p.drawString(340, y, "______________________________")
    y -= 12
    p.drawString(110, y, "ENTREGADO POR")
    p.drawString(375, y, "RECIBIDO POR")

    p.showPage()
    p.save()
    return response



@require_POST
@login_required
def ajustar_stock_conteo(request, conteo_id):
    conteo = get_object_or_404(ConteoInventario, id=conteo_id)
    if not conteo.cerrado:
        messages.error(request, "Debes cerrar el conteo antes de ajustar el stock.")
        return redirect("detalle_conteo", conteo_id=conteo.id)

    items = (
        ItemConteo.objects
        .filter(conteo=conteo)
        .values('producto')
        .annotate(total_contado=Sum('cantidad_contada'))
    )

    cambios = 0
    for item in items:
        producto = Producto.objects.get(id=item["producto"])
        cantidad_nueva = item["total_contado"]
        cantidad_anterior = producto.stock
        diferencia = cantidad_nueva - cantidad_anterior

        # Solo registrar movimiento si hay diferencia
        if diferencia != 0:
            MovimientoInventario.objects.create(
                producto=producto,
                tipo='entrada' if diferencia > 0 else 'salida',
                cantidad=abs(diferencia),
                usuario=request.user
            )

        producto.stock = cantidad_nueva  # Siempre actualiza, aunque no haya diferencia
        producto.save()
        cambios += 1

    messages.success(request, f"Stock actualizado según conteo. {cambios} producto(s) procesado(s).")
    return redirect("detalle_conteo", conteo_id=conteo.id)


def descargar_plantilla_excel(request):
    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Plantilla"

    # Encabezados basados en tu modelo Producto
    encabezados = [
        "Cód. Barras", "Referencia", "Departamento", "Seccion", "Descripción", "Tipo", "Stocks"
    ]

    # Aplicar encabezados con estilo
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num, value=encabezado)
        celda.font = Font(bold=True)

    # Opcional: ejemplo de fila vacía o guía
    ws.append(["123456789", "REF001", "Ropa", "Caballeros", "Camisa Azul", "Unidad", 10])

    # Preparar respuesta HTTP para descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_inventario.xlsx'
    wb.save(response)
    return response
